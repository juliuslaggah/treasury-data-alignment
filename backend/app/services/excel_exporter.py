from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Final

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import ExportError
from app.services.account_matcher import AlignmentResult
from app.services.duplicate_resolver import DuplicateResolution
from app.services.validator import ValidationResult

BALANCE_SOURCE_COLUMN: Final[str] = "DAILY BAL"
BALANCE_DESTINATION_COLUMN: Final[int] = 2

AUDIT_SHEET_NAME: Final[str] = "Alignment Audit"
VALIDATION_SHEET_NAME: Final[str] = "Validation Issues"
UNMATCHED_SOURCE_SHEET_NAME: Final[str] = "Unmatched Source"
DUPLICATE_RESOLUTION_SHEET_NAME: Final[str] = (
    "Duplicate Resolutions"
)

AUDIT_SHEET_NAMES: Final[tuple[str, ...]] = (
    AUDIT_SHEET_NAME,
    VALIDATION_SHEET_NAME,
    UNMATCHED_SOURCE_SHEET_NAME,
    DUPLICATE_RESOLUTION_SHEET_NAME,
)

HEADER_FILL: Final[str] = "1F4E78"
HEADER_FONT_COLOR: Final[str] = "FFFFFF"


class ExcelExporter:
    """Export validated aligned balances into an Excel template."""

    def export(
        self,
        template_path: Path | None = None,
        output_path: Path | None = None,
        alignment: AlignmentResult | None = None,
        validation: ValidationResult | None = None,
        *,
        master_path: Path | None = None,
        excluded_accounts: pd.DataFrame | None = None,
        duplicate_resolutions: tuple[
            DuplicateResolution, ...
        ] = (),
    ) -> Path:
        """
        Export aligned data while preserving the original workbook.

        ``template_path`` remains supported for existing callers.
        ``master_path`` is the preferred name for the approved master
        workbook used by the processing workflow.

        Excluded duplicate rows are assigned a zero balance in the
        generated workbook. The original workbook is never modified.
        """

        selected_template = (
            master_path
            if master_path is not None
            else template_path
        )

        if selected_template is None:
            raise ExportError(
                "An Excel master template path is required."
            )

        if output_path is None:
            raise ExportError(
                "An Excel output path is required."
            )

        if alignment is None:
            raise ExportError(
                "An alignment result is required for export."
            )

        if validation is None:
            raise ExportError(
                "A validation result is required for export."
            )

        template = Path(selected_template)
        output = Path(output_path)

        resolved_excluded_accounts = (
            excluded_accounts.copy(deep=True)
            if excluded_accounts is not None
            else pd.DataFrame()
        )

        self._validate_export_request(
            template=template,
            output=output,
            validation=validation,
        )

        temporary_output = output.with_name(
            f"{output.stem}.temporary{output.suffix}"
        )

        try:
            output.parent.mkdir(
                parents=True,
                exist_ok=True,
            )

            workbook = load_workbook(
                filename=template,
                data_only=False,
                keep_vba=template.suffix.lower() == ".xlsm",
            )

            self._remove_existing_audit_sheets(workbook)

            self._write_aligned_balances(
                workbook,
                alignment,
            )
            self._write_excluded_duplicate_balances(
                workbook,
                resolved_excluded_accounts,
            )
            self._write_alignment_audit(
                workbook,
                alignment,
            )
            self._write_validation_issues(
                workbook,
                validation,
            )
            self._write_unmatched_source(
                workbook,
                alignment,
            )
            self._write_duplicate_resolutions(
                workbook,
                resolved_excluded_accounts,
                duplicate_resolutions,
            )

            workbook.save(temporary_output)
            temporary_output.replace(output)

        except ExportError:
            self._remove_temporary_file(temporary_output)
            raise
        except Exception as exc:
            self._remove_temporary_file(temporary_output)
            raise ExportError(
                f"Unable to export the aligned Excel report: {exc}"
            ) from exc

        return output

    @staticmethod
    def _validate_export_request(
        template: Path,
        output: Path,
        validation: ValidationResult,
    ) -> None:
        """Validate the export request before opening the workbook."""

        if not template.exists():
            raise ExportError(
                f"The Excel template does not exist: {template}"
            )

        if not template.is_file():
            raise ExportError(
                f"The Excel template path is not a file: {template}"
            )

        if template.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ExportError(
                "The template must be an XLSX or XLSM workbook."
            )

        if output.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ExportError(
                "The output file must use the XLSX or XLSM extension."
            )

        if not validation.is_valid:
            error_messages = "; ".join(
                issue.message
                for issue in validation.issues
                if issue.severity == "ERROR"
            )

            raise ExportError(
                "The alignment contains validation errors and cannot "
                f"be exported. {error_messages}"
            )

    def _write_aligned_balances(
        self,
        workbook: Workbook,
        alignment: AlignmentResult,
    ) -> None:
        """Write aligned daily balances into their master rows."""

        if alignment.aligned.empty:
            return

        required_columns = {"MASTER ROW", "MATCH STATUS"}
        missing_columns = required_columns.difference(
            alignment.aligned.columns
        )

        if missing_columns:
            missing = ", ".join(sorted(missing_columns))
            raise ExportError(
                f"Aligned data is missing required columns: {missing}."
            )

        for _, aligned_row in alignment.aligned.iterrows():
            sheet_name = self._resolve_sheet_name(
                workbook,
                aligned_row,
            )
            worksheet = workbook[sheet_name]

            if not isinstance(worksheet, Worksheet):
                raise ExportError(
                    f"Destination is not a worksheet: {sheet_name}"
                )

            master_row = self._to_excel_row(
                aligned_row.get("MASTER ROW")
            )
            match_status = str(
                aligned_row.get("MATCH STATUS", "")
            ).strip().upper()

            balance: int | float = 0

            if match_status == "MATCHED":
                balance = self._to_excel_number(
                    aligned_row.get(BALANCE_SOURCE_COLUMN)
                )

            worksheet.cell(
                row=master_row,
                column=BALANCE_DESTINATION_COLUMN,
                value=balance,
            )

    def _write_excluded_duplicate_balances(
        self,
        workbook: Workbook,
        excluded_accounts: pd.DataFrame,
    ) -> None:
        """Write zero to every deliberately excluded duplicate row."""

        if excluded_accounts.empty:
            return

        if "MASTER ROW" not in excluded_accounts.columns:
            raise ExportError(
                "Excluded duplicate data is missing the "
                "MASTER ROW column."
            )

        for _, excluded_row in excluded_accounts.iterrows():
            sheet_name = self._resolve_sheet_name(
                workbook,
                excluded_row,
            )
            worksheet = workbook[sheet_name]

            if not isinstance(worksheet, Worksheet):
                raise ExportError(
                    f"Destination is not a worksheet: {sheet_name}"
                )

            master_row = self._to_excel_row(
                excluded_row.get("MASTER ROW")
            )

            worksheet.cell(
                row=master_row,
                column=BALANCE_DESTINATION_COLUMN,
                value=0,
            )

    @staticmethod
    def _resolve_sheet_name(
        workbook: Workbook,
        aligned_row: pd.Series[Any],
    ) -> str:
        """Resolve the destination worksheet for a master row."""

        supplied_name = aligned_row.get("SHEET NAME", "")

        if supplied_name is not None and supplied_name is not pd.NA:
            sheet_name = str(supplied_name).strip()

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ExportError(
                        "Template worksheet does not exist: "
                        f"{sheet_name}"
                    )

                return sheet_name

        active_sheet = workbook.active

        if not isinstance(active_sheet, Worksheet):
            raise ExportError(
                "The template does not contain an active worksheet."
            )

        return active_sheet.title

    @staticmethod
    def _to_excel_row(value: object) -> int:
        """Convert a master-row value into a valid Excel row."""

        if value is None or value is pd.NA:
            raise ExportError(
                "An aligned account does not have a master row number."
            )

        try:
            decimal_row = Decimal(str(value))
        except InvalidOperation as exc:
            raise ExportError(
                f"Invalid master row number: {value}"
            ) from exc

        if decimal_row != decimal_row.to_integral_value():
            raise ExportError(
                f"Invalid master row number: {value}"
            )

        row_number = int(decimal_row)

        if row_number < 1:
            raise ExportError(
                f"Invalid master row number: {row_number}"
            )

        return row_number

    @staticmethod
    def _to_excel_number(value: object) -> int | float:
        """Convert a financial value into an Excel-compatible number."""

        if value is None or value is pd.NA:
            return 0

        if isinstance(value, float) and pd.isna(value):
            return 0

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            return value

        if isinstance(value, Decimal):
            decimal_value = value
        else:
            cleaned_value = str(value).strip().replace(",", "")

            if not cleaned_value:
                return 0

            try:
                decimal_value = Decimal(cleaned_value)
            except InvalidOperation as exc:
                raise ExportError(
                    f"Invalid balance value: {value}"
                ) from exc

        if decimal_value == decimal_value.to_integral_value():
            return int(decimal_value)

        return float(decimal_value)

    def _write_alignment_audit(
        self,
        workbook: Workbook,
        alignment: AlignmentResult,
    ) -> None:
        """Create the complete alignment-audit worksheet."""

        worksheet = workbook.create_sheet(AUDIT_SHEET_NAME)
        self._write_dataframe(
            worksheet,
            alignment.aligned,
            "No aligned account records were produced.",
        )

    def _write_validation_issues(
        self,
        workbook: Workbook,
        validation: ValidationResult,
    ) -> None:
        """Create the validation-issues worksheet."""

        worksheet = workbook.create_sheet(
            VALIDATION_SHEET_NAME
        )
        frame = pd.DataFrame(
            [
                {
                    "CODE": issue.code,
                    "SEVERITY": issue.severity,
                    "MESSAGE": issue.message,
                    "ROW COUNT": issue.row_count,
                }
                for issue in validation.issues
            ],
            columns=[
                "CODE",
                "SEVERITY",
                "MESSAGE",
                "ROW COUNT",
            ],
        )

        self._write_dataframe(
            worksheet,
            frame,
            "No validation issues were found.",
        )

    def _write_unmatched_source(
        self,
        workbook: Workbook,
        alignment: AlignmentResult,
    ) -> None:
        """Create the unmatched-source worksheet."""

        worksheet = workbook.create_sheet(
            UNMATCHED_SOURCE_SHEET_NAME
        )
        self._write_dataframe(
            worksheet,
            alignment.unmatched_source,
            "No unmatched source accounts were found.",
        )

    def _write_duplicate_resolutions(
        self,
        workbook: Workbook,
        excluded_accounts: pd.DataFrame,
        resolutions: tuple[DuplicateResolution, ...],
    ) -> None:
        """Create an audit sheet for deliberate duplicate decisions."""

        worksheet = workbook.create_sheet(
            DUPLICATE_RESOLUTION_SHEET_NAME
        )

        resolution_by_key = {
            resolution.account_key: resolution
            for resolution in resolutions
        }

        records: list[dict[str, object]] = []

        if not excluded_accounts.empty:
            for _, excluded_row in excluded_accounts.iterrows():
                account_key = self._clean_text(
                    excluded_row.get("ACCOUNT KEY")
                )
                resolution = resolution_by_key.get(account_key)

                records.append(
                    {
                        "ACCOUNT KEY": account_key,
                        "ACCOUNT NAME": self._clean_text(
                            excluded_row.get("ACCOUNT NAME")
                        ),
                        "KEPT SHEET": (
                            resolution.keep_sheet_name
                            if resolution is not None
                            else ""
                        ),
                        "KEPT MASTER ROW": (
                            resolution.keep_master_row
                            if resolution is not None
                            else None
                        ),
                        "EXCLUDED SHEET": self._clean_text(
                            excluded_row.get("SHEET NAME")
                        ),
                        "EXCLUDED MASTER ROW": (
                            excluded_row.get("MASTER ROW")
                        ),
                        "EXCLUDED SOURCE INDEX": (
                            excluded_row.get("SOURCE INDEX")
                        ),
                        "RESOLUTION STATUS": self._clean_text(
                            excluded_row.get(
                                "RESOLUTION STATUS",
                                "EXCLUDED DUPLICATE",
                            )
                        ),
                        "EXCLUDED BALANCE": 0,
                    }
                )

        frame = pd.DataFrame(
            records,
            columns=[
                "ACCOUNT KEY",
                "ACCOUNT NAME",
                "KEPT SHEET",
                "KEPT MASTER ROW",
                "EXCLUDED SHEET",
                "EXCLUDED MASTER ROW",
                "EXCLUDED SOURCE INDEX",
                "RESOLUTION STATUS",
                "EXCLUDED BALANCE",
            ],
        )

        self._write_dataframe(
            worksheet,
            frame,
            "No duplicate-account resolutions were required.",
        )

    @staticmethod
    def _clean_text(value: object) -> str:
        """Convert an optional table value into clean audit text."""

        if value is None or value is pd.NA:
            return ""

        if isinstance(value, float) and pd.isna(value):
            return ""

        return str(value).strip()

    def _write_dataframe(
        self,
        worksheet: Worksheet,
        frame: pd.DataFrame,
        empty_message: str,
    ) -> None:
        """Write a DataFrame to a professionally formatted sheet."""

        if len(frame.columns) == 0:
            worksheet["A1"] = "STATUS"
            worksheet["A2"] = empty_message
            self._format_audit_sheet(worksheet)
            return

        for column_number, column_name in enumerate(
            frame.columns,
            start=1,
        ):
            worksheet.cell(
                row=1,
                column=column_number,
                value=str(column_name),
            )

        for row_number, row_values in enumerate(
            frame.itertuples(index=False, name=None),
            start=2,
        ):
            for column_number, value in enumerate(
                row_values,
                start=1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                    value=self._to_excel_value(value),
                )

        if frame.empty:
            worksheet.cell(
                row=2,
                column=1,
                value=empty_message,
            )

        self._format_audit_sheet(worksheet)

    @staticmethod
    def _to_excel_value(value: object) -> Any:
        """Convert pandas and Decimal values to Excel-safe values."""

        if value is None or value is pd.NA:
            return None

        if isinstance(value, float) and pd.isna(value):
            return None

        if isinstance(value, Decimal):
            return float(value)

        return value

    @staticmethod
    def _format_audit_sheet(worksheet: Worksheet) -> None:
        """Format an audit worksheet consistently."""

        header_fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL,
        )
        header_font = Font(
            bold=True,
            color=HEADER_FONT_COLOR,
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.freeze_panes = "A2"

        if worksheet.max_column > 0:
            last_header = worksheet.cell(
                row=1,
                column=worksheet.max_column,
            ).coordinate
            worksheet.auto_filter.ref = f"A1:{last_header}"

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            column_letter = get_column_letter(column_number)
            maximum_length = 0

            for row_number in range(
                1,
                worksheet.max_row + 1,
            ):
                cell_value = worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value

                if cell_value is None:
                    continue

                maximum_length = max(
                    maximum_length,
                    len(str(cell_value)),
                )

            worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 12),
                45,
            )

    @staticmethod
    def _remove_existing_audit_sheets(
        workbook: Workbook,
    ) -> None:
        """Remove audit worksheets from an earlier export."""

        for sheet_name in AUDIT_SHEET_NAMES:
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]

            if not isinstance(worksheet, Worksheet):
                raise ExportError(
                    f"Audit sheet is not a worksheet: {sheet_name}"
                )

            workbook.remove(worksheet)

    @staticmethod
    def _remove_temporary_file(path: Path) -> None:
        """Remove an incomplete temporary output."""

        if path.exists():
            path.unlink()


