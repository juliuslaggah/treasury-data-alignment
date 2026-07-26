from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.core.exceptions import ExportError
from app.services.account_matcher import AlignmentResult
from app.services.duplicate_resolver import DuplicateResolution
from app.services.excel_exporter import ExcelExporter
from app.services.validator import (
    ValidationIssue,
    ValidationResult,
)


def create_template(path: Path) -> None:
    """Create a small professional workbook for exporter testing."""

    workbook = Workbook()
    worksheet = workbook.create_sheet("Treasury Report")

    worksheet["A1"] = "ACCOUNT NAME"
    worksheet["B1"] = "DAILY BAL"
    worksheet["C1"] = "CONTROL"

    worksheet["A5"] = "ACCOUNT ONE"
    worksheet["B5"] = 0
    worksheet["C5"] = "=B5*2"

    worksheet["A6"] = "ACCOUNT TWO"
    worksheet["B6"] = 0
    worksheet["C6"] = "=B6*2"

    worksheet["A7"] = "ACCOUNT ONE"
    worksheet["B7"] = 999
    worksheet["C7"] = "=B7*2"

    worksheet["B5"].number_format = "#,##0.00"
    worksheet["B6"].number_format = "#,##0.00"
    worksheet["B7"].number_format = "#,##0.00"

    fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    worksheet["B5"].fill = fill
    worksheet["B6"].fill = fill
    worksheet["B7"].fill = fill

    workbook.save(path)


def create_valid_alignment() -> AlignmentResult:
    """Create a valid alignment containing matched and unmatched rows."""

    aligned = pd.DataFrame(
        [
            {
                "SHEET NAME": "Treasury Report",
                "MASTER ROW": 5,
                "ACCOUNT NAME": "ACCOUNT ONE",
                "MATCH STATUS": "MATCHED",
                "MATCH METHOD": "ACCOUNT NAME",
                "SOURCE INDEX": 0,
                "DAILY BAL": Decimal("1250.50"),
            },
            {
                "SHEET NAME": "Treasury Report",
                "MASTER ROW": 6,
                "ACCOUNT NAME": "ACCOUNT TWO",
                "MATCH STATUS": "UNMATCHED",
                "MATCH METHOD": "",
                "SOURCE INDEX": pd.NA,
                "DAILY BAL": pd.NA,
            },
        ]
    )

    return AlignmentResult(
        aligned=aligned,
        unmatched_source=pd.DataFrame(),
        unmatched_master=pd.DataFrame(
            [
                {
                    "SHEET NAME": "Treasury Report",
                    "MASTER ROW": 6,
                    "ACCOUNT NAME": "ACCOUNT TWO",
                }
            ]
        ),
    )


def test_exports_aligned_balances_into_template(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    create_template(template_path)

    alignment = create_valid_alignment()
    validation = ValidationResult(issues=())

    result_path = ExcelExporter().export(
        template_path=template_path,
        output_path=output_path,
        alignment=alignment,
        validation=validation,
    )

    workbook = load_workbook(
        result_path,
        data_only=False,
    )
    worksheet = workbook["Treasury Report"]

    assert result_path == output_path
    assert worksheet["B5"].value == 1250.5
    assert worksheet["B6"].value == 0
    assert worksheet["B5"].number_format == "#,##0.00"


def test_preserves_template_formulas_and_formatting(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    create_template(template_path)

    ExcelExporter().export(
        template_path=template_path,
        output_path=output_path,
        alignment=create_valid_alignment(),
        validation=ValidationResult(issues=()),
    )

    original = load_workbook(
        template_path,
        data_only=False,
    )
    exported = load_workbook(
        output_path,
        data_only=False,
    )

    original_sheet = original["Treasury Report"]
    exported_sheet = exported["Treasury Report"]

    assert exported_sheet["C5"].value == original_sheet["C5"].value
    assert exported_sheet["C6"].value == original_sheet["C6"].value
    assert (
        exported_sheet["B5"].fill.fgColor.rgb
        == original_sheet["B5"].fill.fgColor.rgb
    )


def test_adds_audit_and_validation_sheets(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    create_template(template_path)

    validation = ValidationResult(
        issues=(
            ValidationIssue(
                code="UNMATCHED_MASTER_ACCOUNT",
                message=(
                    "One master account has no source match."
                ),
                severity="WARNING",
                row_count=1,
            ),
        )
    )

    ExcelExporter().export(
        template_path=template_path,
        output_path=output_path,
        alignment=create_valid_alignment(),
        validation=validation,
    )

    workbook = load_workbook(
        output_path,
        data_only=False,
    )

    assert "Alignment Audit" in workbook.sheetnames
    assert "Validation Issues" in workbook.sheetnames
    assert "Unmatched Source" in workbook.sheetnames
    assert "Duplicate Resolutions" in workbook.sheetnames
    assert (
        workbook["Validation Issues"]["A2"].value
        == "UNMATCHED_MASTER_ACCOUNT"
    )


def test_refuses_to_export_invalid_alignment(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    create_template(template_path)

    validation = ValidationResult(
        issues=(
            ValidationIssue(
                code="UNMATCHED_SOURCE_ACCOUNT",
                message=(
                    "One source account could not be matched."
                ),
                severity="ERROR",
                row_count=1,
            ),
        )
    )

    with pytest.raises(ExportError):
        ExcelExporter().export(
            template_path=template_path,
            output_path=output_path,
            alignment=create_valid_alignment(),
            validation=validation,
        )

    assert output_path.exists() is False


def test_zeros_excluded_duplicate_and_records_resolution(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    create_template(template_path)

    excluded_accounts = pd.DataFrame(
        [
            {
                "ACCOUNT KEY": "ACCOUNT ONE",
                "ACCOUNT NAME": "ACCOUNT ONE",
                "SHEET NAME": "Treasury Report",
                "MASTER ROW": 7,
                "SOURCE INDEX": 6,
                "RESOLUTION STATUS": "EXCLUDED DUPLICATE",
            }
        ]
    )

    resolution = DuplicateResolution(
        account_key="ACCOUNT ONE",
        keep_sheet_name="Treasury Report",
        keep_master_row=5,
    )

    ExcelExporter().export(
        master_path=template_path,
        output_path=output_path,
        alignment=create_valid_alignment(),
        validation=ValidationResult(issues=()),
        excluded_accounts=excluded_accounts,
        duplicate_resolutions=(resolution,),
    )

    original = load_workbook(
        template_path,
        data_only=False,
    )
    exported = load_workbook(
        output_path,
        data_only=False,
    )

    original_sheet = original["Treasury Report"]
    exported_sheet = exported["Treasury Report"]
    audit_sheet = exported["Duplicate Resolutions"]

    assert original_sheet["B7"].value == 999
    assert exported_sheet["B5"].value == 1250.5
    assert exported_sheet["B7"].value == 0
    assert exported_sheet["C7"].value == "=B7*2"
    assert exported_sheet["B7"].number_format == "#,##0.00"

    assert audit_sheet["A2"].value == "ACCOUNT ONE"
    assert audit_sheet["B2"].value == "ACCOUNT ONE"
    assert audit_sheet["C2"].value == "Treasury Report"
    assert audit_sheet["D2"].value == 5
    assert audit_sheet["E2"].value == "Treasury Report"
    assert audit_sheet["F2"].value == 7
    assert audit_sheet["G2"].value == 6
    assert audit_sheet["H2"].value == "EXCLUDED DUPLICATE"
    assert audit_sheet["I2"].value == 0

    